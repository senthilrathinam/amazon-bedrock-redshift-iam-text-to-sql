"""
Excel Knowledge Loader — parses a customer Excel workbook (3 tabs: Tables, Columns, Queries)
and provisions the schema in Redshift with COMMENT ON metadata, relationships, and golden queries.

Expected Excel format:
  Tab 1 "Tables":  columns = [Tables, Description]
  Tab 2 "Columns": columns = [table_name, column_name, data_type, comment]
  Tab 3 "Queries": columns = [User Question, Expected Query]
"""
import os
import yaml
import openpyxl
from typing import Dict, List, Tuple, Optional

EXAMPLES_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "examples.yaml")
RELATIONSHIPS_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "relationships.yaml")

# Redshift data type mapping from common pg types
DTYPE_MAP = {
    "character varying": "VARCHAR(500)",
    "integer": "INTEGER",
    "bigint": "BIGINT",
    "numeric": "NUMERIC(15,4)",
    "boolean": "BOOLEAN",
    "timestamp without time zone": "TIMESTAMP",
    "date": "DATE",
    "text": "TEXT",
    "double precision": "FLOAT",
    "real": "FLOAT",
    "smallint": "SMALLINT",
}


def parse_excel(file_path_or_bytes) -> dict:
    """Parse the 3-tab Excel workbook. Accepts file path (str) or BytesIO object.
    Returns dict with keys: tables, columns, queries."""
    if isinstance(file_path_or_bytes, str):
        wb = openpyxl.load_workbook(file_path_or_bytes)
    else:
        wb = openpyxl.load_workbook(file_path_or_bytes)

    result = {"tables": [], "columns": [], "queries": []}

    # Tab 1: Tables
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            # Clean table name: strip " table" suffix, lowercase, strip whitespace
            name = str(row[0]).strip().lower()
            if name.endswith(" table"):
                name = name[:-6].strip()
            result["tables"].append({
                "table_name": name,
                "description": str(row[1]).strip() if row[1] else ""
            })

    # Tab 2: Columns
    ws = wb.worksheets[1]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            result["columns"].append({
                "table_name": str(row[0]).strip().lower(),
                "column_name": str(row[1]).strip().lower(),
                "data_type": str(row[2]).strip().lower() if row[2] else "character varying",
                "comment": str(row[3]).strip() if row[3] else None
            })

    # Tab 3: Queries
    ws = wb.worksheets[2]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            result["queries"].append({
                "question": str(row[0]).strip(),
                "sql": str(row[1]).strip()
            })

    return result


def _build_ddl(schema: str, parsed: dict) -> List[str]:
    """Build CREATE TABLE DDL statements from parsed Excel data."""
    # Group columns by table
    table_cols: Dict[str, List[dict]] = {}
    for col in parsed["columns"]:
        table_cols.setdefault(col["table_name"], []).append(col)

    ddl_list = [f"CREATE SCHEMA IF NOT EXISTS {schema}"]

    for table_info in parsed["tables"]:
        tbl = table_info["table_name"]
        cols = table_cols.get(tbl, [])
        if not cols:
            continue
        col_defs = []
        for c in cols:
            rs_type = DTYPE_MAP.get(c["data_type"], "VARCHAR(500)")
            col_defs.append(f"    {c['column_name']} {rs_type}")
        ddl = f"CREATE TABLE IF NOT EXISTS {schema}.{tbl} (\n"
        ddl += ",\n".join(col_defs)
        ddl += "\n)"
        ddl_list.append(ddl)

    return ddl_list


def _detect_join_columns(parsed: dict) -> List[Tuple[str, str, str, str]]:
    """Auto-detect JOIN relationships by finding columns with the same name across tables.
    Filters to likely join keys (columns containing 'id', 'number', 'key', 'code')."""
    col_to_tables: Dict[str, List[str]] = {}
    for col in parsed["columns"]:
        col_to_tables.setdefault(col["column_name"], []).append(col["table_name"])

    # Only consider columns that look like join keys
    join_indicators = ('id', 'number', 'key', 'code')

    relationships = []
    for col_name, tables in col_to_tables.items():
        unique_tables = list(set(tables))
        if len(unique_tables) < 2:
            continue
        if not any(ind in col_name.lower() for ind in join_indicators):
            continue
        primary = unique_tables[0]
        for other in unique_tables[1:]:
            relationships.append((other, col_name, primary, col_name))

    return relationships


def save_examples(schema: str, queries: List[dict], parsed_tables: Optional[List[dict]] = None):
    """Save golden queries to examples.yaml, replacing any schema prefix in SQL with target schema."""
    import re as _re

    data = {}
    if os.path.exists(EXAMPLES_PATH):
        with open(EXAMPLES_PATH, 'r') as f:
            data = yaml.safe_load(f) or {}

    # Collect known table names to detect schema prefixes in SQL
    table_names = set()
    if parsed_tables:
        table_names = {t["table_name"] for t in parsed_tables}

    entries = []
    for q in queries:
        sql = q["sql"]
        # Find all schema.table_name patterns where table_name is one of our known tables
        for tbl in table_names:
            pattern = r'(\w+)\.' + _re.escape(tbl) + r'\b'
            for match in _re.finditer(pattern, sql, _re.IGNORECASE):
                old_schema = match.group(1)
                if old_schema.lower() != schema.lower():
                    sql = sql.replace(f"{old_schema}.{tbl}", f"{schema}.{tbl}")
        entries.append({"question": q["question"], "sql": sql})

    data[schema] = entries

    with open(EXAMPLES_PATH, 'w') as f:
        yaml.dump(data, f, default_flow_style=False, sort_keys=False)


def save_relationships(schema: str, rels: List[Tuple[str, str, str, str]]):
    """Save detected relationships to relationships.yaml."""
    data = {}
    if os.path.exists(RELATIONSHIPS_PATH):
        with open(RELATIONSHIPS_PATH, 'r') as f:
            data = yaml.safe_load(f) or {}

    entries = []
    for src_tbl, src_col, tgt_tbl, tgt_col in rels:
        entries.append({
            "source": f"{src_tbl}.{src_col}",
            "target": f"{tgt_tbl}.{tgt_col}",
            "description": f"{src_tbl} references {tgt_tbl} via {src_col}"
        })
    data[schema] = entries

    with open(RELATIONSHIPS_PATH, 'w') as f:
        yaml.dump(data, f, default_flow_style=False, sort_keys=False)


def provision_schema(schema: str, parsed: dict, execute_query_func, load_sample_data: bool = False):
    """Create schema, tables, and apply COMMENT ON metadata in Redshift.
    Optionally loads sample data for testing.
    Returns (success: bool, message: str)."""
    try:
        from .redshift_connector_iam import get_redshift_connection
        conn = get_redshift_connection()
        cur = conn.cursor()

        # Drop and recreate
        cur.execute(f"DROP SCHEMA IF EXISTS {schema} CASCADE")
        conn.commit()

        # Create tables
        for ddl in _build_ddl(schema, parsed):
            cur.execute(ddl)
            conn.commit()

        # Apply table comments
        for t in parsed["tables"]:
            if t["description"]:
                cur.execute(f"COMMENT ON TABLE {schema}.{t['table_name']} IS %s", (t["description"],))
        conn.commit()

        # Apply column comments
        commented = 0
        for c in parsed["columns"]:
            if c["comment"]:
                try:
                    cur.execute(f"COMMENT ON COLUMN {schema}.{c['table_name']}.{c['column_name']} IS %s", (c["comment"],))
                    commented += 1
                except Exception:
                    conn.rollback()
        conn.commit()

        # Load sample data if requested
        data_msg = ""
        if load_sample_data:
            data_msg = _load_sample_data_for_schema(schema, parsed, cur, conn)

        # Save relationships and examples (with schema prefix replacement)
        rels = _detect_join_columns(parsed)
        save_relationships(schema, rels)
        save_examples(schema, parsed["queries"], parsed["tables"])

        conn.close()

        table_count = len(parsed["tables"])
        query_count = len(parsed["queries"])
        msg = f"Created {table_count} tables, {commented} column comments, {len(rels)} relationships, {query_count} golden queries."
        if data_msg:
            msg += f" {data_msg}"
        return True, msg

    except Exception as e:
        return False, f"Error: {str(e)}"


def _load_sample_data_for_schema(schema: str, parsed: dict, cur, conn) -> str:
    """Generate and insert sample data. Uses mortgage-specific generators if tables match,
    otherwise generates generic placeholder data."""
    table_names = {t["table_name"] for t in parsed["tables"]}
    known_tables = {"origination_currentversion", "originationborrower_currentversion", "originationproperty_currentversion"}

    if table_names == known_tables:
        return _load_mortgage_sample_data(schema, cur, conn)

    # Generic sample data for unknown schemas
    import random
    from datetime import datetime, timedelta
    random.seed(42)
    total_rows = 0

    table_cols: Dict[str, List[dict]] = {}
    for c in parsed["columns"]:
        table_cols.setdefault(c["table_name"], []).append(c)

    for tbl, cols in table_cols.items():
        rows = []
        for i in range(20):
            row = []
            for c in cols:
                dt = c["data_type"]
                if "int" in dt or "bigint" in dt:
                    row.append(random.randint(1, 1000))
                elif "numeric" in dt or "float" in dt or "double" in dt or "real" in dt:
                    row.append(round(random.uniform(1, 100000), 2))
                elif "bool" in dt:
                    row.append(random.choice([True, False]))
                elif "timestamp" in dt or "date" in dt:
                    row.append(datetime(2025, 1, 1) + timedelta(days=random.randint(0, 365)))
                else:
                    row.append(f"sample_{c['column_name']}_{i}")
            rows.append(tuple(row))

        placeholders = ",".join(["%s"] * len(cols))
        col_names = ",".join(c["column_name"] for c in cols)
        sql = f"INSERT INTO {schema}.{tbl} ({col_names}) VALUES ({placeholders})"
        for row in rows:
            cur.execute(sql, row)
        conn.commit()
        total_rows += len(rows)

    return f"Loaded {total_rows} generic sample rows."


def _load_mortgage_sample_data(schema: str, cur, conn) -> str:
    """Load mortgage-specific sample data using the genai_poc bootstrapper generators."""
    import random
    random.seed(42)
    from .genai_poc_bootstrapper import (
        generate_origination_data, generate_borrower_data, generate_property_data,
        ORIGINATION_COLS, BORROWER_COLS, PROPERTY_COLS
    )

    def _insert(table, columns, rows):
        placeholders = ",".join(["%s"] * len(columns))
        cols_str = ",".join(columns)
        sql = f"INSERT INTO {schema}.{table} ({cols_str}) VALUES ({placeholders})"
        for i in range(0, len(rows), 50):
            cur.executemany(sql, rows[i:i+50])

    orig_rows = generate_origination_data(100)
    _insert("origination_currentversion", ORIGINATION_COLS, orig_rows)
    conn.commit()

    borr_rows = generate_borrower_data(100)
    _insert("originationborrower_currentversion", BORROWER_COLS, borr_rows)
    conn.commit()

    prop_rows = generate_property_data(100)
    _insert("originationproperty_currentversion", PROPERTY_COLS, prop_rows)
    conn.commit()

    return f"Loaded {len(orig_rows)} origination, {len(borr_rows)} borrower, {len(prop_rows)} property records."
